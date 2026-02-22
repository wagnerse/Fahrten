import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Freie Touren BB+MV 01-03.04.2026"

# Headers
headers = ["Tour-Nr", "Prio", "Tag", "Datum", "Ab", "Startbahnhof", "An", "Zielbahnhof", "Fahrten", "Punkte", "Dauer", "Euro"]
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

# Data - only April 1-3 tours (excluding near-future 22.02 tours)
tours = [
    # Mi 01.04.2026
    [704213, 1, "Mi", "01.04.2026", "03:53", "Lübbenau(Spreewald)", "06:53", "Dessau Hbf", 1, 6, "03:00", 57.00],
    [705313, 1, "Mi", "01.04.2026", "04:10", "Senftenberg", "04:41", "Cottbus Hbf", 1, 8, "00:31", 14.30],
    [704218, 1, "Mi", "01.04.2026", "05:00", "Bad Belzig", "05:43", "Berlin-Wannsee", 1, 8, "00:43", 14.30],
    [705925, 1, "Mi", "01.04.2026", "05:30", "Graal-Müritz", "06:18", "Rostock Hbf", 1, 5, "00:48", 17.45],
    [704219, 1, "Mi", "01.04.2026", "05:33", "Rostock Hbf", "06:21", "Graal-Müritz", 1, 5, "00:48", 17.45],
    [705826, 1, "Mi", "01.04.2026", "05:48", "Warnemünde", "06:44", "Güstrow", 1, 7, "00:56", 17.45],
    [705312, 1, "Mi", "01.04.2026", "05:51", "Cottbus Hbf", "06:18", "Senftenberg", 1, 9, "00:27", 11.30],
    [704347, 1, "Mi", "01.04.2026", "06:02", "Cottbus Hbf", "09:55", "Falkenberg(Elster)", 2, 2, "03:53", 69.35],
    [705978, 1, "Mi", "01.04.2026", "06:30", "Graal-Müritz", "09:18", "Rostock Hbf", 3, 0, "02:48", 45.60],
    [704220, 1, "Mi", "01.04.2026", "06:33", "Rostock Hbf", "07:21", "Graal-Müritz", 1, 0, "00:48", 17.45],
    [705926, 1, "Mi", "01.04.2026", "07:30", "Graal-Müritz", "08:18", "Rostock Hbf", 1, 0, "00:48", 17.45],
    [704349, 1, "Mi", "01.04.2026", "08:01", "Barth", "09:53", "Barth", 2, 0, "01:52", 28.60],
    [705825, 1, "Mi", "01.04.2026", "08:03", "Bad Kleinen", "08:50", "Rostock Hbf", 1, 0, "00:47", 17.45],
    [704221, 1, "Mi", "01.04.2026", "08:33", "Rostock Hbf", "09:21", "Graal-Müritz", 1, 0, "00:48", 17.45],
    [705927, 1, "Mi", "01.04.2026", "09:30", "Graal-Müritz", "10:18", "Rostock Hbf", 1, 0, "00:48", 17.45],
    [704222, 1, "Mi", "01.04.2026", "09:33", "Rostock Hbf", "10:21", "Graal-Müritz", 1, 0, "00:48", 17.45],
    [705311, 1, "Mi", "01.04.2026", "10:00", "Bad Belzig", "10:43", "Berlin-Wannsee", 1, 3, "00:43", 14.30],
    [704344, 1, "Mi", "01.04.2026", "12:29", "Warnemünde", "13:48", "Warnemünde", 2, 2, "01:19", 18.44],
    [704345, 1, "Mi", "01.04.2026", "13:19", "Warnemünde", "14:58", "Warnemünde", 2, 2, "01:39", 23.10],
    [706037, 1, "Mi", "01.04.2026", "17:01", "Bad Kleinen", "17:56", "Lübeck Hbf", 1, 0, "00:55", 17.45],
    [705472, 1, "Mi", "01.04.2026", "17:18", "Ueckermünde Stadthafen", "17:57", "Pasewalk", 1, 3, "00:39", 14.30],
    [704208, 1, "Mi", "01.04.2026", "17:39", "Warnemünde", "18:50", "Güstrow", 1, 1, "01:11", 22.48],
    [705371, 1, "Mi", "01.04.2026", "18:02", "Lübeck Hbf", "18:51", "Bad Kleinen", 1, 0, "00:49", 17.45],
    [705305, 1, "Mi", "01.04.2026", "18:45", "Jüterbog", "19:24", "Falkenberg(Elster)", 1, 3, "00:39", 14.30],
    [705877, 1, "Mi", "01.04.2026", "19:00", "Barth", "21:37", "Stralsund Hbf", 3, 0, "02:37", 42.90],
    [705302, 1, "Mi", "01.04.2026", "19:26", "Szczecin Glowny", "21:21", "Angermünde", 1, 0, "01:55", 36.42],
    [704215, 1, "Mi", "01.04.2026", "20:14", "Senftenberg", "23:19", "Bad Belzig", 1, 6, "03:05", 58.58],
    [704203, 1, "Mi", "01.04.2026", "21:31", "Seebad Heringsdorf", "21:40", "Swinoujscie Centrum", 1, 4, "00:09", 11.30],
    # Do 02.04.2026
    [704241, 1, "Do", "02.04.2026", "03:49", "Neustrelitz Hbf", "05:23", "Berlin Südkreuz", 1, 6, "01:34", 29.77],
    [705932, 1, "Do", "02.04.2026", "04:00", "Rostock Hbf", "05:56", "Lübeck Hbf", 1, 6, "01:56", 36.73],
    [705393, 1, "Do", "02.04.2026", "04:02", "Cottbus Hbf", "06:59", "Cottbus Hbf", 2, 7, "02:57", 42.78],
    [705390, 1, "Do", "02.04.2026", "04:04", "Ruhland", "07:52", "Rathenow", 2, 6, "03:48", 67.48],
    [705378, 1, "Do", "02.04.2026", "04:19", "Elsterwerda", "04:22", "Elsterwerda-Biehla", 1, 9, "00:03", 11.30],
    [705928, 1, "Do", "02.04.2026", "04:31", "Wolgast", "05:40", "Swinoujscie Centrum", 1, 4, "01:09", 21.85],
    [704249, 1, "Do", "02.04.2026", "04:59", "Cottbus Hbf", "05:54", "Falkenberg(Elster)", 1, 7, "00:55", 17.45],
    [704232, 1, "Do", "02.04.2026", "05:02", "Jüterbog", "07:13", "Stendal Hbf", 1, 4, "02:11", 41.48],
    [704238, 1, "Do", "02.04.2026", "05:57", "Dresden-Neustadt", "07:43", "Cottbus Hbf", 1, 6, "01:46", 33.57],
    [704237, 1, "Do", "02.04.2026", "06:26", "Schwerin Hbf", "07:24", "Rostock Hbf", 1, 2, "00:58", 17.45],
    [705383, 1, "Do", "02.04.2026", "07:00", "Bad Belzig", "07:43", "Berlin-Wannsee", 1, 3, "00:43", 14.30],
    [704246, 1, "Do", "02.04.2026", "08:50", "Cottbus Hbf", "09:54", "Elsterwerda", 1, 1, "01:04", 20.27],
    [704250, 1, "Do", "02.04.2026", "09:00", "Bad Belzig", "09:43", "Berlin-Wannsee", 1, 3, "00:43", 14.30],
    [706038, 1, "Do", "02.04.2026", "09:08", "Rostock Hbf", "09:54", "Bad Kleinen", 1, 0, "00:46", 17.45],
    [705384, 1, "Do", "02.04.2026", "09:53", "Cottbus Hbf", "10:18", "Senftenberg", 1, 4, "00:25", 11.30],
    [704247, 1, "Do", "02.04.2026", "10:50", "Cottbus Hbf", "11:50", "Elsterwerda-Biehla", 1, 2, "01:00", 19.00],
    [705385, 1, "Do", "02.04.2026", "12:12", "Falkenberg(Elster)", "13:25", "Hoyerswerda", 1, 1, "01:13", 23.12],
    [704253, 1, "Do", "02.04.2026", "12:33", "Rostock Hbf", "13:21", "Graal-Müritz", 1, 0, "00:48", 17.45],
    [704245, 1, "Do", "02.04.2026", "13:35", "Potsdam Hbf", "14:40", "Berlin Gesundbrunnen", 1, 1, "01:05", 20.58],
    [704248, 1, "Do", "02.04.2026", "14:12", "Falkenberg(Elster)", "15:25", "Hoyerswerda", 1, 1, "01:13", 23.12],
    [704323, 1, "Do", "02.04.2026", "15:30", "Graal-Müritz", "17:31", "Schwerin Hbf", 2, 1, "02:01", 34.90],
    [704236, 1, "Do", "02.04.2026", "16:15", "Stralsund Hbf", "16:55", "Barth", 1, 3, "00:40", 14.30],
    [704322, 1, "Do", "02.04.2026", "16:30", "Graal-Müritz", "18:21", "Graal-Müritz", 2, 0, "01:51", 30.40],
    [704227, 1, "Do", "02.04.2026", "16:33", "Hoyerswerda", "18:52", "Leipzig Hbf", 1, 1, "02:19", 44.02],
    [705931, 1, "Do", "02.04.2026", "16:33", "Rostock Hbf", "17:21", "Graal-Müritz", 1, 0, "00:48", 17.45],
    [704251, 1, "Do", "02.04.2026", "17:00", "Bad Belzig", "17:43", "Berlin-Wannsee", 1, 3, "00:43", 14.30],
    [705979, 1, "Do", "02.04.2026", "18:30", "Graal-Müritz", "22:00", "Rostock Hbf", 3, 4, "03:30", 48.99],
    [704373, 1, "Do", "02.04.2026", "20:14", "Cottbus Hbf", "01:15", "Ruhland", 3, 8, "05:01", 76.83],
    [704332, 1, "Do", "02.04.2026", "20:47", "Stendal Hbf", "00:28", "Falkenberg(Elster)", 2, 6, "03:41", 54.52],
    [704254, 1, "Do", "02.04.2026", "21:15", "Rostock Hbf", "22:03", "Graal-Müritz", 1, 5, "00:48", 17.45],
    [705930, 1, "Do", "02.04.2026", "22:02", "Lübeck Hbf", "23:56", "Rostock Hbf", 1, 4, "01:54", 36.10],
    [704244, 1, "Do", "02.04.2026", "22:26", "Senftenberg", "01:31", "Bad Belzig", 1, 6, "03:05", 58.58],
    [704231, 1, "Do", "02.04.2026", "23:02", "Lübeck Hbf", "00:16", "Schwerin Hbf", 1, 4, "01:14", 23.43],
    # Fr 03.04.2026
    [704272, 1, "Fr", "03.04.2026", "06:09", "Cottbus Hbf", "06:29", "Forst(Lausitz)", 1, 7, "00:20", 12.43],
    [704264, 2, "Fr", "03.04.2026", "06:12", "Stralsund Hbf", "08:25", "Angermünde", 1, 2, "02:13", 46.33],
    [704260, 1, "Fr", "03.04.2026", "08:08", "Züssow", "09:40", "Swinoujscie Centrum", 1, 2, "01:32", 32.04],
    [704273, 1, "Fr", "03.04.2026", "08:30", "Forst(Lausitz)", "08:49", "Cottbus Hbf", 1, 7, "00:19", 12.43],
    [704276, 1, "Fr", "03.04.2026", "08:33", "Hoyerswerda", "10:52", "Leipzig Hbf", 1, 4, "02:19", 48.42],
    [705933, 1, "Fr", "03.04.2026", "10:30", "Graal-Müritz", "11:18", "Rostock Hbf", 1, 3, "00:48", 19.20],
    [704281, 1, "Fr", "03.04.2026", "10:57", "Dresden-Neustadt", "12:25", "Hoyerswerda", 1, 4, "01:28", 30.66],
    [704274, 1, "Fr", "03.04.2026", "11:30", "Forst(Lausitz)", "11:49", "Cottbus Hbf", 1, 7, "00:19", 12.43],
    [704288, 1, "Fr", "03.04.2026", "11:30", "Graal-Müritz", "12:18", "Rostock Hbf", 1, 3, "00:48", 19.20],
    [704287, 1, "Fr", "03.04.2026", "11:33", "Rostock Hbf", "12:21", "Graal-Müritz", 1, 3, "00:48", 19.20],
    [705980, 1, "Fr", "03.04.2026", "12:30", "Graal-Müritz", "15:18", "Rostock Hbf", 3, 2, "02:48", 50.16],
    [705934, 1, "Fr", "03.04.2026", "12:33", "Rostock Hbf", "13:21", "Graal-Müritz", 1, 3, "00:48", 19.20],
    [704261, 1, "Fr", "03.04.2026", "15:19", "Zinnowitz", "15:55", "Seebad Heringsdorf", 1, 4, "00:36", 15.73],
    [704278, 1, "Fr", "03.04.2026", "16:26", "Schwerin Hbf", "17:24", "Rostock Hbf", 1, 5, "00:58", 19.20],
    [704279, 1, "Fr", "03.04.2026", "19:09", "Güstrow", "20:24", "Warnemünde", 1, 4, "01:15", 26.13],
    [704277, 1, "Fr", "03.04.2026", "20:55", "Falkenberg(Elster)", "21:56", "Cottbus Hbf", 1, 4, "01:01", 21.25],
    [704280, 1, "Fr", "03.04.2026", "23:19", "Elsterwerda", "23:22", "Elsterwerda-Biehla", 1, 12, "00:03", 12.43],
]

# Day color fills
day_fills = {
    "Mi": PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),  # light blue
    "Do": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),  # light green
    "Fr": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),  # light orange
}

for row_idx, tour in enumerate(tours, 2):
    day = tour[2]
    fill = day_fills.get(day)
    for col_idx, value in enumerate(tour, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        if fill:
            cell.fill = fill

# Column widths
widths = [10, 6, 5, 12, 6, 25, 6, 25, 8, 7, 7, 10]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

# Euro format
for row in range(2, len(tours) + 2):
    cell = ws.cell(row=row, column=12)
    cell.number_format = '#,##0.00 €'

# Freeze header
ws.freeze_panes = 'A2'

# Auto filter
ws.auto_filter.ref = f"A1:L{len(tours) + 1}"

wb.save("/Users/wagnerse/Desktop/Fahrten/Freie_Touren_BB_MV_01-03_April_2026.xlsx")
print(f"Excel erstellt mit {len(tours)} Touren")
